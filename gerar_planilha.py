"""
Gera uma planilha XLSX com os atributos que cada projeto de extensão deve ter.
Baseado nos campos definidos em data/extensoes.yaml (RF18).
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
import yaml
import os

# ── Cores (paleta UFRJ / azul-marinho + verde) ─────────────────────────────
COR_CABECALHO     = "1A3A5C"   # azul-marinho escuro
COR_OBRIG_FUNDO   = "E8F0FA"   # azul muito claro (obrigatório)
COR_OPCIO_FUNDO   = "F0FAF0"   # verde muito claro (opcional)
COR_TIPO_FUNDO    = "FFF8E1"   # amarelo claro (coluna Tipo)
COR_BORDER        = "B0BEC5"

COLUNAS = [
    # (chave_yaml, Nome para exibir, Tipo, Obrigatório, Descrição, Exemplo)
    ("titulo",             "Título",                "Texto",      True,
     "Nome completo do projeto",
     "Olimpíada Brasileira de Informática — Polo IC"),

    ("coordenador",        "Coordenador",           "Texto",      True,
     "Nome do(a) docente responsável",
     "Hugo Nobrega"),

    ("descricao",          "Descrição",             "Texto longo",True,
     "Resumo do projeto (3-5 linhas)",
     "O IC atua como polo de aplicação das provas..."),

    ("vagas",              "Vagas",                 "Inteiro",    True,
     "Número de vagas disponíveis para alunos",
     "5"),

    ("bolsa",              "Bolsa?",                "Booleano",   True,
     "Indica se há bolsa remunerada (true/false)",
     "false"),

    ("processo_seletivo",  "Processo Seletivo",     "Texto",      True,
     "Como o aluno deve se candidatar",
     "Análise de currículo pelos coordenadores"),

    ("perfil",             "Perfil do Aluno",       "Texto",      True,
     "Pré-requisitos e características desejadas no candidato",
     "Alunos com interesse em competições de programação"),

    ("contato",            "Contato",               "E-mail",     True,
     "E-mail de contato do projeto",
     "obi@ic.ufrj.br"),

    # ── Opcionais ───────────────────────────────────────────────────────────
    ("area",               "Área",                  "Texto",      False,
     "Grande área temática do projeto",
     "Educação e Divulgação Científica"),

    ("modalidade",         "Modalidade",            "Texto",      False,
     "Tipo: Projeto, Programa, Curso, Evento…",
     "Projeto"),

    ("carga_horaria",      "Carga Horária (h/sem)", "Inteiro",    False,
     "Horas semanais de dedicação esperadas",
     "4"),

    ("link_inscricao",     "Link de Inscrição",     "URL",        False,
     "URL externa de inscrição (se houver)",
     "https://olimpiada.ic.unicamp.br/"),

    ("ano_inicio",         "Ano de Início",         "Inteiro",    False,
     "Ano em que o projeto foi iniciado",
     "2023"),

    ("ano_fim",            "Ano de Fim",            "Inteiro",    False,
     "Ano de encerramento (omitir se ainda ativo)",
     "2025"),
]

CABECALHOS = ["Campo (YAML)", "Nome Exibido", "Tipo", "Obrigatório?",
              "Descrição", "Exemplo"]


def cor_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def borda_fina() -> Border:
    lado = Side(style="thin", color=COR_BORDER)
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def build_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()

    # ── Aba 1: Atributos ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Atributos dos Projetos"

    # Título principal (linha 1, mesclada)
    ws.merge_cells("A1:F1")
    celula_titulo = ws["A1"]
    celula_titulo.value = "Atributos dos Projetos de Extensão — IC/UFRJ"
    celula_titulo.font = Font(name="Calibri", bold=True, size=14,
                              color="FFFFFF")
    celula_titulo.fill = cor_fill(COR_CABECALHO)
    celula_titulo.alignment = Alignment(horizontal="center", vertical="center",
                                        wrap_text=True)
    ws.row_dimensions[1].height = 32

    # Subtítulo (linha 2)
    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value = ("Campos obrigatórios (RF18): titulo, coordenador, descricao, "
                 "vagas, bolsa, processo_seletivo, perfil, contato   |   "
                 "Campos opcionais: area, modalidade, carga_horaria, "
                 "link_inscricao, ano_inicio, ano_fim")
    sub.font = Font(name="Calibri", italic=True, size=9, color="555555")
    sub.alignment = Alignment(horizontal="center", vertical="center",
                              wrap_text=True)
    ws.row_dimensions[2].height = 28

    # Cabeçalho das colunas (linha 3)
    for col_idx, nome in enumerate(CABECALHOS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=nome)
        cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        cell.fill = cor_fill(COR_CABECALHO)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = borda_fina()
    ws.row_dimensions[3].height = 20

    # Dados
    for row_idx, (chave, nome, tipo, obrig, desc, exemplo) in \
            enumerate(COLUNAS, start=4):
        obrig_str = "✔ Obrigatório" if obrig else "Opcional"
        valores = [chave, nome, tipo, obrig_str, desc, exemplo]
        fundo = COR_OBRIG_FUNDO if obrig else COR_OPCIO_FUNDO

        for col_idx, val in enumerate(valores, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = borda_fina()

            # Fundo especial para coluna "Obrigatório?"
            if col_idx == 4:
                cell.font = Font(name="Calibri", size=10, bold=True,
                                 color="1A3A5C" if obrig else "2E7D32")
                cell.fill = cor_fill("D6E4F7" if obrig else "C8E6C9")
            else:
                cell.fill = cor_fill(fundo)

        ws.row_dimensions[row_idx].height = 40

    # Larguras das colunas
    larguras = [22, 26, 16, 17, 48, 52]
    for col_idx, larg in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = larg

    # Congelar painel (fixar linhas de título + cabeçalho)
    ws.freeze_panes = "A4"

    # ── Aba 2: Projetos (template em branco para preenchimento) ─────────────
    ws2 = wb.create_sheet("Projetos — Preencher")

    ws2.merge_cells("A1:N1")
    cab2 = ws2["A1"]
    cab2.value = "Cadastro de Projetos de Extensão — IC/UFRJ  (preencha a partir da linha 3)"
    cab2.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    cab2.fill = cor_fill(COR_CABECALHO)
    cab2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    colunas_template = [
        ("titulo",            "Título",                "obrig"),
        ("coordenador",       "Coordenador",           "obrig"),
        ("descricao",         "Descrição",             "obrig"),
        ("vagas",             "Vagas",                 "obrig"),
        ("bolsa",             "Bolsa? (true/false)",   "obrig"),
        ("processo_seletivo", "Processo Seletivo",     "obrig"),
        ("perfil",            "Perfil do Aluno",       "obrig"),
        ("contato",           "Contato (e-mail)",      "obrig"),
        ("area",              "Área",                  "opcio"),
        ("modalidade",        "Modalidade",            "opcio"),
        ("carga_horaria",     "Carga Horária (h/sem)", "opcio"),
        ("link_inscricao",    "Link de Inscrição",     "opcio"),
        ("ano_inicio",        "Ano de Início",         "opcio"),
        ("ano_fim",           "Ano de Fim",            "opcio"),
    ]

    for col_idx, (chave, nome, tipo) in enumerate(colunas_template, start=1):
        fundo = COR_OBRIG_FUNDO if tipo == "obrig" else COR_OPCIO_FUNDO
        cell_nome = ws2.cell(row=2, column=col_idx, value=nome)
        cell_nome.font = Font(name="Calibri", bold=True, size=10,
                              color="FFFFFF")
        cell_nome.fill = cor_fill(
            "1A3A5C" if tipo == "obrig" else "2E7D32"
        )
        cell_nome.alignment = Alignment(horizontal="center", vertical="center",
                                        wrap_text=True)
        cell_nome.border = borda_fina()

        # 10 linhas em branco para preenchimento
        for row_idx in range(3, 13):
            c = ws2.cell(row=row_idx, column=col_idx, value="")
            c.fill = cor_fill(fundo)
            c.border = borda_fina()
            c.alignment = Alignment(vertical="top", wrap_text=True)

        ws2.column_dimensions[get_column_letter(col_idx)].width = 22

    ws2.row_dimensions[2].height = 28
    for r in range(3, 13):
        ws2.row_dimensions[r].height = 45
    ws2.freeze_panes = "A3"

    # ── Legenda (rodapé na aba 1) ───────────────────────────────────────────
    linha_leg = len(COLUNAS) + 5
    ws.merge_cells(f"A{linha_leg}:F{linha_leg}")
    leg = ws[f"A{linha_leg}"]
    leg.value = ("🔵 Fundo azul claro = campo obrigatório (RF18)   |   "
                 "🟢 Fundo verde claro = campo opcional")
    leg.font = Font(name="Calibri", italic=True, size=9, color="333333")
    leg.alignment = Alignment(horizontal="center")

    return wb


if __name__ == "__main__":
    caminho_saida = os.path.join(
        os.path.dirname(__file__),
        "atributos_projetos_extensao.xlsx"
    )
    wb = build_workbook()
    wb.save(caminho_saida)
    print(f"✅ Planilha gerada com sucesso: {caminho_saida}")
