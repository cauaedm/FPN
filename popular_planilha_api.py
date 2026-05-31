"""
Popula a planilha XLSX com todas as ações de extensão da UFRJ
vindas da API pública: https://portal.extensao.ufrj.br/php/listaAcoes.php
"""

import json
import urllib.request
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

URL_API = "https://portal.extensao.ufrj.br/php/listaAcoes.php"

# ── Paleta de cores ─────────────────────────────────────────────────────────
COR_CABECALHO   = "1A3A5C"
COR_IC          = "E8F0FA"   # linha do IC (azul claro)
COR_OUTROS      = "FFFFFF"   # outras unidades
COR_IC_ALT      = "D0E4F5"   # linhas pares do IC
COR_OUTROS_ALT  = "F5F5F5"   # linhas pares de outros

COLUNAS = [
    ("id_anuncio_acao",         "ID Anúncio"),
    ("centro",                  "Centro"),
    ("unidade",                 "Unidade"),
    ("coordenador",             "Coordenador"),
    ("titulo",                  "Título"),
    ("modalidade",              "Modalidade"),
    ("area_principal",          "Área Principal"),
    ("area_secundaria",         "Área Secundária"),
    ("formato_realizacao",      "Formato"),
    ("carga_horaria",           "Carga Horária (h)"),
    ("data_inicio",             "Início"),
    ("data_termino",            "Término"),
    ("data_inicio_inscricoes",  "Início Inscrições"),
    ("data_termino_inscricoes", "Término Inscrições"),
    ("como_inscrever",          "Como se Inscrever"),
    ("link_inscricoes",         "Link de Inscrição"),
    ("publico",                 "Público-Alvo"),
    ("contato",                 "Contato"),
    ("email_atendimento",       "E-mail Atendimento"),
    ("telefone_atendimento",    "Telefone"),
    ("resumo",                  "Resumo"),
    ("descricao",               "Descrição"),
]


def cor_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def borda():
    lado = Side(style="thin", color="B0BEC5")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def eh_ic(acao):
    unidade = (acao.get("unidade") or "").lower()
    return "computa" in unidade


print(f"Buscando dados de: {URL_API}")
req = urllib.request.Request(URL_API, headers={"User-Agent": "Mozilla/5.0"})
with urllib.request.urlopen(req, timeout=30) as resp:
    dados = json.loads(resp.read().decode("utf-8"))

print(f"Total de acoes recebidas: {len(dados)}")
acoes_ic    = [a for a in dados if eh_ic(a)]
acoes_outras = [a for a in dados if not eh_ic(a)]
print(f"  -> IC: {len(acoes_ic)}  |  Outras unidades: {len(acoes_outras)}")

# ── Ordenar: IC primeiro, depois resto ──────────────────────────────────────
acoes_ordenadas = acoes_ic + acoes_outras

wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════════════════════
# ABA 1 – Todas as ações
# ════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Todas as Extensões"

# Título
ws.merge_cells(f"A1:{get_column_letter(len(COLUNAS))}1")
c = ws["A1"]
c.value = f"Ações de Extensão UFRJ — Fonte: {URL_API}  |  Total: {len(dados)} ações  |  IC: {len(acoes_ic)}"
c.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
c.fill = cor_fill(COR_CABECALHO)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 26

# Cabeçalhos
for col_i, (_, nome) in enumerate(COLUNAS, start=1):
    cell = ws.cell(row=2, column=col_i, value=nome)
    cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    cell.fill = cor_fill(COR_CABECALHO)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = borda()
ws.row_dimensions[2].height = 22

# Dados
for row_i, acao in enumerate(acoes_ordenadas, start=3):
    is_ic  = eh_ic(acao)
    par    = (row_i % 2 == 0)
    if is_ic:
        fundo = COR_IC_ALT if par else COR_IC
    else:
        fundo = COR_OUTROS_ALT if par else COR_OUTROS

    for col_i, (chave, _) in enumerate(COLUNAS, start=1):
        val = acao.get(chave, "")
        # Listas (ex: faixa_etaria_publico) → join
        if isinstance(val, list):
            val = "; ".join(str(v) for v in val)
        cell = ws.cell(row=row_i, column=col_i, value=val)
        cell.font = Font(name="Calibri", size=9,
                         bold=is_ic and col_i in (5,))   # título em negrito p/ IC
        cell.fill = cor_fill(fundo)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = borda()

    ws.row_dimensions[row_i].height = 40

# Larguras aproximadas
larguras = [8, 30, 28, 24, 40, 14, 22, 22, 16, 12, 12, 12, 14, 14, 24, 32, 30, 22, 22, 14, 60, 60]
for i, w in enumerate(larguras, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUNAS))}{len(acoes_ordenadas)+2}"

# ════════════════════════════════════════════════════════════════════════════
# ABA 2 – Somente IC
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Somente IC")

ws2.merge_cells(f"A1:{get_column_letter(len(COLUNAS))}1")
c2 = ws2["A1"]
c2.value = f"Ações de Extensão — Instituto de Computação UFRJ  |  Total: {len(acoes_ic)} ações"
c2.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
c2.fill = cor_fill(COR_CABECALHO)
c2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 26

for col_i, (_, nome) in enumerate(COLUNAS, start=1):
    cell = ws2.cell(row=2, column=col_i, value=nome)
    cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    cell.fill = cor_fill(COR_CABECALHO)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = borda()
ws2.row_dimensions[2].height = 22

for row_i, acao in enumerate(acoes_ic, start=3):
    par   = (row_i % 2 == 0)
    fundo = COR_IC_ALT if par else COR_IC
    for col_i, (chave, _) in enumerate(COLUNAS, start=1):
        val = acao.get(chave, "")
        if isinstance(val, list):
            val = "; ".join(str(v) for v in val)
        cell = ws2.cell(row=row_i, column=col_i, value=val)
        cell.font = Font(name="Calibri", size=9)
        cell.fill = cor_fill(fundo)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = borda()
    ws2.row_dimensions[row_i].height = 40

for i, w in enumerate(larguras, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.freeze_panes = "A3"
if acoes_ic:
    ws2.auto_filter.ref = f"A2:{get_column_letter(len(COLUNAS))}{len(acoes_ic)+2}"

# ── Salvar ───────────────────────────────────────────────────────────────────
caminho = os.path.join(os.path.dirname(__file__), "extensoes_ufrj_api.xlsx")
wb.save(caminho)
print(f"Planilha gerada: {caminho}")
